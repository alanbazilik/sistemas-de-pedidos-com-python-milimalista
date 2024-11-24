from flask import Flask, jsonify, request, render_template, send_from_directory
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import os

app = Flask(__name__)

# Configuração do diretório de uploads
UPLOAD_FOLDER = 'uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Dados dos itens
items = [
    {"id": 1, "name": "Hambúrguer", "price": 15.0, "image": "images/Hambúrguer.jpg"},
    {"id": 2, "name": "Pizza", "price": 30.0, "image": "images/Pizza.jpg"},
    {"id": 3, "name": "Refrigerante", "price": 5.0, "image": "images/Refrigerante.png"},
]

# Lista de pedidos
orders = []

# Criar diretório de uploads caso não exista
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

# Rota principal para carregar o front-end
@app.route("/")
def index():
    return render_template("index.html")

# Endpoint para obter os itens
@app.route("/items", methods=["GET"])
def get_items():
    return jsonify(items)

# Endpoint para adicionar itens ao pedido
@app.route("/add_to_order", methods=["POST"])
def add_to_order():
    data = request.json
    item_id = data.get("id")
    item = next((i for i in items if i["id"] == item_id), None)
    if item:
        orders.append(item)
        return jsonify({"message": f"{item['name']} adicionado ao pedido!"}), 200
    return jsonify({"error": "Item não encontrado"}), 404

# Endpoint para gerar a comanda no Excel
@app.route("/generate_excel", methods=["POST"])
def generate_excel():
    data = request.json
    customer_name = data.get("customer")
    order_items = data.get("orders")

    if not customer_name:
        return jsonify({"error": "Nome do cliente não informado!"}), 400

    if not order_items:
        return jsonify({"error": "Nenhum item no pedido!"}), 400

    # Criar Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Comanda"

    # Configuração de estilos
    header_font = Font(bold=True, size=12, color="FFFFFF")
    regular_font = Font(size=11)
    center_alignment = Alignment(horizontal="center", vertical="center")
    left_alignment = Alignment(horizontal="left", vertical="center")
    border_style = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_fill = PatternFill("solid", fgColor="4CAF50")
    total_fill = PatternFill("solid", fgColor="FFEB3B")

    # Títulos
    ws.merge_cells("A1:B1")
    ws["A1"] = "Comanda do Cliente"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = center_alignment

    ws["A2"] = "Cliente:"
    ws["A2"].font = header_font
    ws["B2"] = customer_name
    ws["B2"].font = regular_font
    ws["B2"].alignment = left_alignment

    # Cabeçalho
    ws.append(["Item", "Preço"])
    for cell in ws[4]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border_style
        cell.alignment = center_alignment

    # Adicionar itens do pedido
    total = 0
    row = 5
    for item in order_items:
        ws.append([item["name"], item["price"]])
        ws[f"A{row}"].alignment = left_alignment
        ws[f"B{row}"].alignment = center_alignment
        ws[f"A{row}"].border = border_style
        ws[f"B{row}"].border = border_style
        total += item["price"]
        row += 1

    # Linha Total
    ws.append(["", ""])
    ws.append(["Total:", total])
    total_row = row + 1
    ws[f"A{total_row}"].font = Font(bold=True, size=12)
    ws[f"A{total_row}"].alignment = center_alignment
    ws[f"B{total_row}"].font = Font(bold=True, size=12)
    ws[f"B{total_row}"].alignment = center_alignment
    ws[f"B{total_row}"].fill = total_fill

    # Ajustar largura das colunas
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15

    # Salvar Excel no diretório de uploads
    filename = f"{customer_name.replace(' ', '_')}_comanda.xlsx"
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    wb.save(filepath)

    return jsonify({
        "message": f"Comanda gerada com sucesso para {customer_name}!",
        "file": filename
    }), 200

# Rota para servir os arquivos gerados
@app.route('/uploads/<filename>')
def download_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

if __name__ == "__main__":
    app.run(debug=True)
